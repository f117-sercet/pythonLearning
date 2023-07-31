# 获取单元格中的指定数据

import os
import openpyxl


#方法1：指定坐标的方式
#sheet[“A1”]

path =r"C:\Users\LENOVO\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook.active


cell1=sheet['A1']
print(cell1.value)



#方法二：指定行列的方式 sheet.cell(row=, column=)方式
cell1 = sheet.cell(row=1,column=1)         # 获取第1行第1列的数据
cell2 = sheet.cell(row=3,column=2)         # 获取第3行第4的数据
print(cell1.value,cell2.value)


