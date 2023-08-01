import os
import openpyxl

path = r"C:\Users\asuka\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')	# 返回一个workbook数据类型的值
print(workbook.sheetnames)	# 打印Excel表中的所有表

path = r"C:\Users\asuka\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  # 返回一个workbook数据类型的值
print(workbook.sheetnames)  # 打印Excel表中的所有表

sheet = workbook['Sheet1']  # 获取指定sheet表
print(sheet)


path = r"C:\Users\asuka\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook.active     # 获取活动表
print(sheet)

path = r"C:\Users\asuka\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook['Sheet1']  # 获取指定sheet表
print(sheet.dimensions)     # 获取表格的尺寸大小




