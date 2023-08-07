from random import random
import openpyxl
import os

## 获取文件路径
path =r"C:\Users\LENOVO\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('jobInfo.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook.active

print(sheet.dimensions)


## 获取相对应的列名

dataColumns={}
for i in sheet.iter_rows(min_row=1,max_row=1,min_col=2):
    for j in i:
        dataColumn={}
        #dataColumn.update(j.value,j.value)
        dataColumn[j.value]=j.value
        dataColumns.update(dataColumn)

print(dataColumns)
print(dataColumns.__getitem__('FK_MDNB_MATCHING_CODE'))

## 获取指定的列的数据

maxColumn = sheet.max_row
for i in sheet.iter_rows(min_row=2,max_row=maxColumn,min_col=2):
   for j in i:
       columnData={}





## 连接数据库，然后更新数据

# 获取指定列的数据






